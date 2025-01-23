import os
import re
import unicodedata
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment

from get_DB_data import get_stdid_data, get_fiscal_data, get_sale_data, get_fin_ind_data

'''
【檢查 1】(twn.push.p_u11_now)的主旨(subject)年季和內文(txt)的迄日年季要相同
'''
# 國字數字 -> 阿拉伯數字，e.g. 公告本公司一一二年第一季合併財務報告業經董事會通過 -> 公告本公司112年第1季合併財務報告業經董事會通過
def chinese_to_arabic(subject):
    chinese_numerals = {
        "零": "0", 
        "一": "1", "壹": "1", "ㄧ": "1", 
        "二": "2", "貳": "2",
        "三": "3", "參": "3",
        "四": "4", "肆": "4",
        "五": "5", "伍": "5",
        "六": "6", "陸": "6",
        "七": "7", "柒": "7",
        "八": "8", "捌": "8", 
        "九": "9", "玖": "9"
    }
    
    letter_list = []
    for letter in subject:
        if letter in chinese_numerals:
            letter_list.append(chinese_numerals[letter])
        else:
            letter_list.append(letter)
    
    arabic_subject = ''.join(letter_list)
    return arabic_subject 


def check_same_YQ(extract, df):
    
    rule_1_violation_idx = []
    
    for idx in range(len(extract)):
        subject = unicodedata.normalize('NFC', df['subject'][idx])  # 年年，兩個字編碼不同，標準化
        subject_ = subject.replace(' ', '').replace('百', '').replace('十', '')
        subject_ = chinese_to_arabic(subject_)
        
        year_pattern = r"(?<![\(\d])(\d{3,4})(?![\)\d])"  # 113, 2024, 公告慧友電子(5484)董事會決議通過民國111年度合併財務報告
        match = re.search(year_pattern, subject_)
        if match:
            subject_year = next((g for g in match.groups() if g), None)
            if len(subject_year) == 3:
                subject_year = str(int(subject_year) + 1911)
        else:
            print(idx, f'A: 年度不存在於主旨: {subject}')
            rule_1_violation_idx.append((idx, f'A: 年度不存在於主旨: {subject}'))
            continue
        
        quarter_pattern = r"[第前]?(\d)季|Q(\d)"  # 第2季, 前3季, 3季, Q4
        match = re.search(quarter_pattern, subject_)
        if match: 
            subject_quarter = next((g for g in match.groups() if g), None)
            subject_quarter = 'Q' + subject_quarter
        elif '上半年' in subject_:  # 上半年度 or 上半年
            subject_quarter = 'Q2'
        elif '年度' in subject_:
            subject_quarter = 'Q4'
        elif '年度' not in subject_ and '季度' not in subject_ and '度' in subject_:  
            # 公告本公司董事會通過112年上半度合併財務報告
            if '半度' in subject_:    
                subject_quarter = 'Q2'
            # 公告本公司董事會通過111度合併財務報告
            else:
                subject_quarter = 'Q4'
        else:
            Q4_pattern = r"年(?:合併|個別|)(?:財務報告|財務報表|財報)"  # 兩個年編碼不同
            match = re.search(Q4_pattern, subject_)
            if match:
                subject_quarter = 'Q4'
            else:
                date_pattern = r"[~～-]\s*(\d{2,4})[年./-]?(\d{1,2})[月./-]?(\d{1,2})[日./]?"
                match = re.search(date_pattern, subject_)
                if match:
                    subject_month = match.group(2)
                    subject_quarter = str(int(int(subject_month) / 3))
                    subject_quarter = 'Q' + subject_quarter
                else:
                    print(idx, f'A: 季度不存在於主旨: {subject}')
                    rule_1_violation_idx.append((idx, f'A: 季度不存在於主旨: {subject}'))
                    continue
            
        subject_YQ = subject_year + subject_quarter
            
        fin_end_date = extract['財務迄日'][idx]  # YYYYMMDD
        if pd.isna(fin_end_date):
            print(idx, 'A: 財務迄日不存在無法判斷')
            rule_1_violation_idx.append((idx, 'A: 財務迄日不存在無法判斷'))
            continue
        else:
            fin_end_date_year = fin_end_date[:4]
            fin_end_date_quarter = fin_end_date[4:6]
            if fin_end_date_quarter not in ['03', '06', '09', '12']:
                print(idx, f'A: 財務迄日不在 3, 6, 9, 12 月: {fin_end_date}')
                rule_1_violation_idx.append((idx, f'A: 財務迄日不在 3, 6, 9, 12 月: {fin_end_date}'))
                continue
            else:    
                fin_end_date_quarter = 'Q' + str(int(int(fin_end_date[4:6]) / 3))
                fin_end_date_YQ = fin_end_date_year + fin_end_date_quarter
        
        if subject_YQ != fin_end_date_YQ:
            print(idx, f'A: 主旨年季與財務迄日年季不同。主旨: {subject}, 財務迄日: {fin_end_date}')
            rule_1_violation_idx.append((idx, f'A: 主旨年季與財務迄日年季不同。主旨: {subject}, 財務迄日: {fin_end_date}'))

    return rule_1_violation_idx

# ==================================================================================================================
'''
【檢查 2】檢查日期欄位不為NULL或空白
以下科目不能為NULL或空白：
發言日期、董事會決議日、財務起日、財務迄日
'''

def check_null_date(extract):
    
    rule_2_violation_idx = []
    
    for idx in range(len(extract)):
        reals_date = extract['發言日期'][idx]
        board_date = extract['董事會決議日'][idx]
        fin_beg_date = extract['財務起日'][idx]
        fin_end_date = extract['財務迄日'][idx]    
    
        if pd.isna(reals_date):
            print(idx, 'B: 發言日期不存在')
            rule_2_violation_idx.append((idx, 'B: 發言日期不存在'))
            continue
        if pd.isna(board_date):
            print(idx, 'B: 董事會決議日不存在')
            rule_2_violation_idx.append((idx, 'B: 董事會決議日不存在'))
            continue
        if pd.isna(fin_beg_date):
            print(idx, 'B: 財務起日不存在')
            rule_2_violation_idx.append((idx, 'B: 財務起日不存在'))
            continue
        if pd.isna(fin_end_date):
            print(idx, 'B: 財務迄日不存在')
            rule_2_violation_idx.append((idx, 'B: 財務迄日不存在'))
    
    return rule_2_violation_idx

# ==================================================================================================================
'''
【檢查 3】 檢查日期欄位格式正確性(前提是檢查2要通過)及規則
3.1 u11_board.FIN_BEG_DATE格式
	-起日日期一定要為01/01，其餘則不通過
3.2	u11_board.FIN_END_DATE格式
	-月份為份為1、3、5、7、8、10、12時，迄日月份結束日= 31
	-月份為4、6、 9、11時，迄日月份結束日= 30
	-月份為2時，迄日月份結束日= 28 或 29
3.3	u11_board.AUDIT_COMMITTEE_DATE格式
	假設AUDIT_COMMITTEE_DATE截字截到「不適用」、「無」、「NA」可通過檢查
    -> 沒有「不適用」、「無」、「NA」的文字或是日期格式錯誤不通過(e.g. '12/31')
'''

def check_date_correct(extract, extract_data):
    
    rule_3_violation_idx = []
    
    for idx in range(len(extract)):
        fin_beg_date = extract['財務起日'][idx]
        fin_end_date = extract['財務迄日'][idx]
        audit_committee_date = extract_data['審計委員會通過日'][idx]
        
        # 3.1
        if not pd.isna(fin_beg_date):
            if fin_beg_date[4:] != '0101':
                print(idx, f'C - 3.1: 財務起日不在 1 月 1 日: {fin_beg_date}')
                rule_3_violation_idx.append((idx, f'C - 3.1: 財務起日不在 1 月 1 日: {fin_beg_date}'))
                continue
        else:  # fin_beg_date 空值已在檢查 2 判定
            continue
        
        # 3.2
        if not pd.isna(fin_end_date):
            month, day = fin_end_date[-4:-2], fin_end_date[-2:]
        else:  # fin_end_date 空值已在檢查 2 判定
            continue
        
        if month in ['03', '12']:
            if day != '31':
                print(idx, f'C - 3.2: 財務迄日不在月底: {fin_end_date}')
                rule_3_violation_idx.append((idx, f'C - 3.2: 財務迄日不在月底: {fin_end_date}'))
                continue
        elif month in ['06', '09']:
            if day != '30':
                print(idx, f'C - 3.2: 財務迄日不在月底: {fin_end_date}')
                rule_3_violation_idx.append((idx, f'C - 3.2: 財務迄日不在月底: {fin_end_date}'))
                continue
        else:  # 非3、6、9、12月情況已在檢查 1 判定
            continue
        
        # 3.3 判斷有截到審委員會通過日但不合理的狀況
        if not pd.isna(audit_committee_date) and not re.fullmatch(r"^[\d年月日./-]+$", audit_committee_date):
            if not any(keyword in audit_committee_date for keyword in ['不適用', 'NA', '無']):
                print(idx, f'C - 3.3: 審計委員會通過日不合理: {audit_committee_date}')
                rule_3_violation_idx.append((idx, f'C - 3.3: 審計委員會通過日不合理: {audit_committee_date}')) 
        elif not pd.isna(audit_committee_date) and re.fullmatch(r"^[\d年月日./-]+$", audit_committee_date):
            if len(audit_committee_date) < 6:
                print(idx, f'C - 3.3: 審計委員會通過日不合理: {audit_committee_date}')
                rule_3_violation_idx.append((idx, f'C - 3.3: 審計委員會通過日不合理: {audit_committee_date}')) 
                
    return rule_3_violation_idx

# ==================================================================================================================
'''
【檢查 4】 月制規則
月制取法: 取u11_board.COMP_ID 和 stdid.COMP_ID 配, 得TEJ_COMP_ID ，再用TEJ_COMP_ID和event_fiscal_month.TEJ_COMP_ID配，
限制條件是event_fiscal_month.FIN_END_DATE = ‘2999-12-31’，得到fiscal_month 。
通過檢查條件為fiscal_month = 12，不通過ERR = D
'''

def month_regulation_check(extract, stdid_df, fiscal_df):
    
    # 檢查 4 判斷
    rule_4_violation_idx = []
    
    for idx in range(len(extract)):
        comp_id = extract['公司碼'][idx]
        tej_comp_id = stdid_df[stdid_df['comp_id'] == str(comp_id)]['tej_comp_id'].iloc[0]
        fiscal_month = fiscal_df[fiscal_df['tej_comp_id'] == tej_comp_id]['fiscal_month']
        
        if fiscal_month.iloc[0] != 12:
            print(idx, f'D: 月制 fiscal_month 不等於 12: {fiscal_month}')
            rule_4_violation_idx.append((idx, f'D: 月制 fiscal_month 不等於 12: {fiscal_month}'))
                
    return rule_4_violation_idx

# ==================================================================================================================
'''
【檢查 5】
檢查u11_board.FIN_BEG_DATE、u11_board.FIN_END_DATE、u11_board.REALS_DATE、
u11_board.AUDIT_COMMITTEE_DATE、u11_board.BOARD_DATE彼此之間的關係
(不為NULL且不為空白時檢查)
5.1 u11_board.FIN_BEG_DATE  <  u11_board.FIN_END_DATE
5.2	u11_board.FIN_END_DATE  <  u11_board.BOARD_DATE
5.3	u11_board.REALS_DATE  >  u11_board.FIN_END_DATE
5.4	u11_board.REALS_DATE  >=  u11_board. BOARD_DATE
5.5	u11_board.BOARD_DATE  >=  u11_board.AUDIT_COMMITTEE_DATE
5.6	系統年 >=  u11_board.FIN_END_DATE 的年轉西元後  >=  系統年 - 1年		
5.7 若5.1、5.2、5.3、5.4、5.5、5.6、任一不過 -> ERR = E
'''

def check_date_relation(extract):
    
    rule_5_violation_idx = []
    
    for idx in range(len(extract)):
        reals_date = extract['發言日期'][idx].strftime('%Y%m%d')
        board_date = extract['董事會決議日'][idx]
        audit_committee_date = extract['審計委員會通過日'][idx]
        fin_beg_date = extract['財務起日'][idx]
        fin_end_date = extract['財務迄日'][idx]
        
        if pd.isna([reals_date, board_date, audit_committee_date, fin_beg_date, fin_end_date]).any():
            continue
        
        # 5.1
        if int(fin_beg_date) >= int(fin_end_date):
            print(idx, f'E - 5.1: 財務起日 >= 財務迄日: {fin_beg_date}, {fin_end_date}') 
            rule_5_violation_idx.append((idx, f'E - 5.1: 財務起日 >= 財務迄日: {fin_beg_date}, {fin_end_date}'))
            continue
        
        # 5.2
        if int(fin_end_date) >= int(board_date):
            print(idx, f'E - 5.2: 財務迄日 >= 董事會決議日: {fin_end_date}, {board_date}') 
            rule_5_violation_idx.append((idx, f'E - 5.2: 財務迄日 >= 董事會決議日: {fin_end_date}, {board_date}'))
            continue
            
        # 5.3
        if int(reals_date) <= int(fin_end_date):
            print(idx, f'E - 5.3: 發言日期 <= 財務迄日: {reals_date}, {fin_end_date}') 
            rule_5_violation_idx.append((idx, f'E - 5.3: 發言日期 <= 財務迄日: {reals_date}, {fin_end_date}'))
            continue
        
        # 5.4
        if int(reals_date) < int(board_date):
            print(idx, f'E - 5.4: 發言日期 < 董事會決議日: {reals_date}, {board_date}') 
            rule_5_violation_idx.append((idx, f'E - 5.4: 發言日期 < 董事會決議日: {reals_date}, {board_date}'))
            continue
        
        # 5.5
        if int(board_date) < int(audit_committee_date):
            print(idx, f'E - 5.5: 董事會決議日 < 審計委員會通過日: {board_date}, {audit_committee_date}') 
            rule_5_violation_idx.append((idx, f'E - 5.5: 董事會決議日 < 審計委員會通過日: {board_date}, {audit_committee_date}'))
            continue
        
        # 5.6
        current_year = datetime.now().year
        fin_end_year = int(fin_end_date[:4])
        if not (current_year-1) <= fin_end_year <= current_year:
            print(idx, f'E - 5.6: 財務迄日年 > 系統年或 < 系統年 - 1: {fin_end_year}, {current_year}')
            rule_5_violation_idx.append((idx, f'E - 5.6: 財務迄日年 > 系統年或 < 系統年 - 1: {fin_end_year}, {current_year}'))
            
    return rule_5_violation_idx

# ==================================================================================================================
'''
【檢查 6】 數值欄位格式
6.1	除了欄位u11_board.T3990有小數，其餘在*_txt欄位不得有小數
6.2  u11_board.T3990欄位一定要有小數位
6.3	txt裡有「(」、「（」、「-」、「-%」時，要檢查對應的數值是否 < 0
6.4	u11_board.T3990 和u11_board.T3950的正負號要一致
6.5  6.1、6.2、6.3、6.4任一沒過 -> ERR = F
'''

def check_num(extract):
    
    # 6.1 已在 handle_no_decimal_num() 解決
    
    rule_6_violation_idx = []
    
    for idx in range(len(extract)):
        t_3990 = extract['基本每股盈餘'][idx]
        t_3950 = extract['歸屬於母公司業主淨利'][idx]
        
        # 6.2 
        if pd.isna(t_3990):
            continue
        elif '.' not in str(t_3990):
            print(idx, f'F - 6.2: 基本每股盈餘無小數點: {str(t_3990)}')
            rule_6_violation_idx.append((idx, f'F - 6.2: 基本每股盈餘無小數點: {str(t_3990)}'))
            continue
        
        # 6.4
        if pd.isna(t_3950):
            continue
        elif t_3990 * t_3950 < 0:
            print(idx, f'F - 6.4: 基本每股盈餘與歸屬於母公司業主淨利正負號不一致: {str(t_3990)}, {str(t_3950)}')
            rule_6_violation_idx.append((idx, f'F - 6.4: 基本每股盈餘與歸屬於母公司業主淨利正負號不一致: {str(t_3990)}, {str(t_3950)}'))
            
    return rule_6_violation_idx

# ==================================================================================================================
'''
【檢查 7】 檢查營收和月營收數值差異過大
7.1 u11_board 和 sale連結規則：
BYM = estm_u11_board.FIN_END_DATE(迄日)的年月
FYM =sale.ZYYMM的年月
取estm_u11_board.COMP_ID 和 basic.attr_stdid.COMP_ID 配, 得TEJ_COMP_ID，
再用TEJ_COMP_ID 、BYM和sale.TEJ_COMP_ID、FYM配，去找對應的T8104 且T8104不為NULL且不等於0。
* sale.T8104 IS NULL AND u11_board.FIN_END_DATE >= stdid.FST_LIST_DATE（無月營收資料 且 財務迄日 >= 首次上市日期）-> 5330

7.2 檢查
取u11_board.T3100和T8104做比較，設DIF = ABS[((T3100-T8104))/ABS(T8104)]×100，
若DIF >= 10 則不通過檢查，ERR = G
'''

def check_revenue_difference(extract, stdid_df, sale_df):
    
    rule_7_violation_idx = []
    
    for idx in range(len(extract)):
        
        # 7.1 u11_board 和 sale_df 連結，取得 t_8104
        comp_id = extract['公司碼'][idx]
        fin_end_date = extract['財務迄日'][idx]
        if pd.isna(fin_end_date):
            continue
        BYM = fin_end_date[:6]
        
        comp_df = stdid_df[stdid_df['comp_id'] == str(comp_id)]
        tej_comp_id = comp_df['tej_comp_id'].iloc[0]
        fst_list_date = comp_df['fst_list_date'].iloc[0]
        
        t_8104_row = sale_df[(sale_df['tej_comp_id'] == tej_comp_id) & (sale_df['zyymm'] == BYM)]['t8104']
        
        if fst_list_date != 'None':
            if t_8104_row.empty and int(fin_end_date) >= int(fst_list_date):
                print(idx, f'G - 7.1: 公司碼: {comp_id} 於 {BYM} 無 t_8104 且財務迄日: {fin_end_date} > 首次上市日期: {fst_list_date}')
                rule_7_violation_idx.append((idx, f'G - 7.1: 公司碼: {comp_id} 於 {BYM} 無 t_8104 且財務迄日: {fin_end_date} > 首次上市日期: {fst_list_date}'))
                continue
            elif not t_8104_row.empty and pd.isna(t_8104_row.iloc[0]) and int(fin_end_date) >= int(fst_list_date):
                print(idx, f'G - 7.1: 公司碼: {comp_id} 於 {BYM} t_8104 為 NULL 且財務迄日: {fin_end_date} > 首次上市日期: {fst_list_date}')
                rule_7_violation_idx.append((idx, f'G - 7.1: 公司碼: {comp_id} 於 {BYM} t_8104 為 NULL 且財務迄日: {fin_end_date} > 首次上市日期: {fst_list_date}'))
                continue
            elif (t_8104_row.empty or (not t_8104_row.empty and pd.isna(t_8104_row.iloc[0]))) & (int(fin_end_date) < int(fst_list_date)): 
                continue
        elif t_8104_row.empty or (not t_8104_row.empty and pd.isna(t_8104_row.iloc[0])):
            continue
        
        t_8104 = t_8104_row.iloc[0]
        
        if t_8104 == 0:
            continue
        
        # 7.2 檢查 DIF
        t_3100 = extract['營業收入'][idx]
        if pd.isna(t_3100) or t_3100 == 0:
            continue
        
        DIF = abs((t_3100-t_8104)/abs(t_8104))*100
        if DIF >= 10:
            print(idx, f'G - 7.2: DIF >= 10%: 營業收入: {t_3100}, 月營收: {t_8104}, DIF: {round(DIF, 2)}%')
            rule_7_violation_idx.append((idx, f'G - 7.2: DIF >= 10%: 營業收入: {t_3100}, 月營收: {t_8104}, DIF: {round(DIF, 2)}%'))
            
    return rule_7_violation_idx

# ==================================================================================================================
'''
【檢查 8】檢查科目缺值(BY 產業)
8.1	取u11_board.COMP_ID 和 stdid.COMP_ID 配, 得TEJ_COMP_ID，
    再和event_fin_ind.FIN_IND的TEJ_COMP_ID且event_fin_ind.FIN_END_DATE = ’2999-12-31’ 配，得到對應的FIN_IND

8.2 以下為各產業科目碼不能為NULL的規則
    S_list = ['營業收入', '營業利益', '稅前淨利', '本期淨利', '歸屬於母公司業主淨利', '基本每股盈餘']
    HO_list = ['淨收益_銀行業', '利息淨收益_銀行業', '稅前淨利', '本期淨利', '歸屬於母公司業主淨利', '基本每股盈餘']
    B_list = ['營業收入', '營業利益', '稅前淨利', '本期淨利', '歸屬於母公司業主淨利', '基本每股盈餘']
    F_list = ['營業收入', '營業毛利', '營業利益', '稅前淨利', '本期淨利', '歸屬於母公司業主淨利', '基本每股盈餘']
8.3  8.2任一產業科目碼有缺 -> ERR = H

'''

def check_fin_ind_null(extract, stdid_df, fin_ind_df):
    
    rule_8_violation_idx = []
    
    S_list = ['營業收入', '營業利益', '稅前淨利', '本期淨利', '歸屬於母公司業主淨利', '基本每股盈餘']
    HO_list = ['淨收益_銀行業', '利息淨收益_銀行業', '稅前淨利', '本期淨利', '歸屬於母公司業主淨利', '基本每股盈餘']
    B_list = ['營業收入', '營業利益', '稅前淨利', '本期淨利', '歸屬於母公司業主淨利', '基本每股盈餘']
    F_list = ['營業收入', '營業毛利', '營業利益', '稅前淨利', '本期淨利', '歸屬於母公司業主淨利', '基本每股盈餘']

    for idx in range(len(extract)):
        comp_id = extract['公司碼'][idx]
        tej_comp_id = stdid_df[stdid_df['comp_id'] == str(comp_id)]['tej_comp_id'].iloc[0]
        fin_ind = fin_ind_df[fin_ind_df['tej_comp_id'] == tej_comp_id]['fin_ind'].iloc[0]
        
        if fin_ind == 'F' and any(pd.isna(extract.loc[idx, col]) for col in F_list):
            print(idx, 'H - 8.3: 一般產業存在空值欄位')
            rule_8_violation_idx.append((idx, 'H - 8.3: 一般產業存在空值欄位'))
        elif fin_ind == 'S' and any(pd.isna(extract.loc[idx, col]) for col in S_list):
            print(idx, 'H - 8.3: 證券業存在空值欄位')
            rule_8_violation_idx.append((idx, 'H - 8.3: 證券業存在空值欄位'))
        elif fin_ind in ['H', 'O'] and any(pd.isna(extract.loc[idx, col]) for col in HO_list):
            print(idx, 'H - 8.3: 銀行業 or 金控業存在空值欄位')
            rule_8_violation_idx.append((idx, 'H - 8.3: 銀行業 or 金控業存在空值欄位'))
        elif fin_ind == 'B' and any(pd.isna(extract.loc[idx, col]) for col in B_list):
            print(idx, 'H - 8.3: 保險業存在空值欄位')
            rule_8_violation_idx.append((idx, 'H - 8.3: 保險業存在空值欄位'))
    
    return rule_8_violation_idx

# ==================================================================================================================
'''
【檢查 9】 t0010、t1000、t200e檢查
9.1 不可以為0或NULL
 	u11_board.t0010、u11_board.t1000、u11_board.t200e
9.2	皆要 > 0 
	u11_board.t0010、u11_board.t1000
9.3  9.1+9.2任一不通過 ERR = I
'''

def check_asset_debt_equity(extract):

    rule_9_violation_idx = []
    
    for idx in range(len(extract)):
        
        t_0010 = extract['期末總資產'][idx]
        t_1000 = extract['期末總負債'][idx]
        t_200e = extract['歸屬於母公司權益'][idx]
        
        # 9.1
        if t_0010 == 0 or pd.isna(t_0010):
            print(idx, 'I - 9.1: 期末總資產為 0 或 NULL')
            rule_9_violation_idx.append((idx, 'I - 9.1: 期末總資產為 0 或 NULL'))
            continue
        if t_1000 == 0 or pd.isna(t_1000):
            print(idx, 'I - 9.1: 期末總負債為 0 或 NULL')
            rule_9_violation_idx.append((idx, 'I - 9.1: 期末總負債為 0 或 NULL'))
            continue
        if t_200e == 0 or pd.isna(t_200e):
            print(idx, 'I - 9.1: 歸屬於母公司權益為 0 或 NULL')
            rule_9_violation_idx.append((idx, 'I - 9.1: 歸屬於母公司權益為 0 或 NULL'))
            continue
        
        # 9.2
        if t_0010 < 0:
            print(idx, 'I - 9.2: 期末總資產 < 0')
            rule_9_violation_idx.append((idx, 'I - 9.2: 期末總資產 < 0'))
            continue
        if t_1000 < 0:
            print(idx, 'I - 9.2: 期末總負債 < 0')
            rule_9_violation_idx.append((idx, 'I - 9.2: 期末總負債 < 0'))

    return rule_9_violation_idx

# ==================================================================================================================
'''
【檢查10】 u11_board筆數問題
u11_board只能有唯一的comp_id和reals_date，
假設有多筆相同comp_id和reals_date的資料時，則全部筆數皆進到5330
'''

def check_same_id_reals_date(extract):
    rule_10_violation_idx = extract[extract.duplicated(subset=['公司碼', '發言日期'], keep=False)].index.to_list()
    rule_10_violation_idx = [(idx, 'J: 公司碼與發言日期重複') for idx in rule_10_violation_idx]
    return rule_10_violation_idx

# ==================================================================================================================
def produce_5330(extract, extract_data, df):

    stdid_df = get_stdid_data()
    fiscal_df = get_fiscal_data()
    sale_df = get_sale_data()
    fin_ind_df = get_fin_ind_data()
    
    rule_1_violation_idx = check_same_YQ(extract, df)
    rule_2_violation_idx = check_null_date(extract)
    rule_3_violation_idx = check_date_correct(extract, extract_data)
    rule_4_violation_idx = month_regulation_check(extract, stdid_df, fiscal_df)
    rule_5_violation_idx = check_date_relation(extract)
    rule_6_violation_idx = check_num(extract)
    rule_7_violation_idx = check_revenue_difference(extract, stdid_df, sale_df)
    rule_8_violation_idx = check_fin_ind_null(extract, stdid_df, fin_ind_df)
    rule_9_violation_idx = check_asset_debt_equity(extract)
    rule_10_violation_idx = check_same_id_reals_date(extract)

    violation_idx_list = [
        rule_1_violation_idx, 
        rule_2_violation_idx,
        rule_3_violation_idx,
        rule_4_violation_idx,
        rule_5_violation_idx,
        rule_6_violation_idx,
        rule_7_violation_idx,
        rule_8_violation_idx,
        rule_9_violation_idx,
        rule_10_violation_idx
    ]

    # 整理違規數據到字典 all_violations
    all_violations = {}
    for violation_idx in violation_idx_list:
        for idx, message in violation_idx:
            if idx not in all_violations:
                all_violations[idx] = []
            all_violations[idx].append(message)
            
    # 新增 data_5330，並加入'錯誤代碼與資訊'欄位
    data_5330 = extract.loc[all_violations.keys()].copy()
    data_5330['錯誤代碼與資訊'] = data_5330.index.map(
        lambda idx: "\n".join(all_violations[idx])
    )

    data_5330['發言日期'] = data_5330['發言日期'].astype(str).str.replace('-', '')

    def AD_to_ROC(date):
        
        if pd.isna(date):
            return date
        
        year, month, day = date[:-4], date[-4:-2], date[-2:]
        year_ROC = str(int(year) - 1911)
                
        return f"{year_ROC}{month}{day}"

    date_col_ = ['發言日期', '董事會決議日', '審計委員會通過日', '財務起日', '財務迄日']

    data_5330[date_col_] = data_5330[date_col_].map(AD_to_ROC)

    data_5330 = data_5330.sort_values(
        by=['公司碼', '發言日期', '則次', '合併M/個別A'],
        ascending=[True, True, True, True],
        key=lambda col: col if col.name != '合併M/個別A' else col.map({'M': 0, 'A': 1})
    )

    data_5330['幣別'] = 'TWD'

    no_pass_col = [
        '公司碼', '發言日期', '則次', '合併M/個別A', '董事會決議日', '審計委員會通過日', '財務起日', '財務迄日', '幣別',
        '營業收入', '淨收益_銀行業', '利息淨收益_銀行業', '營業收益_證券業', '營業毛利', '營業利益', '稅前淨利',
        '本期淨利', '歸屬於母公司業主淨利', '基本每股盈餘', '期末總資產', '期末總負債', '歸屬於母公司權益', '其他應敘明事項',
        '營業收入_TXT', '淨收益_銀行業_TXT', '利息淨收益_銀行業_TXT', '營業收益_證券業_TXT', '營業毛利_TXT',
        '營業利益_TXT', '稅前淨利_TXT', '本期淨利_TXT', '歸屬於母公司業主淨利_TXT', '基本每股盈餘_TXT', 
        '期末總資產_TXT', '期末總負債_TXT', '歸屬於母公司權益_TXT', '截字fin_type', '檔名', '錯誤代碼與資訊'
    ]

    data_5330 = data_5330[no_pass_col]
    data_5330.rename(columns={'利息淨收益_銀行業_TXT': '淨利息淨收益_銀行業_TXT'}, inplace=True)  # 更改欄位名稱以符合 5330 表頭
    
    return data_5330

# ==================================================================================================================
def data_5330_to_excel(data_5330):

    # 報表日期設定
    report_date = datetime.today().strftime('%Y%m%d')
    current_time = datetime.now().strftime("%H%M%S")  # 當前時間的 HH:MM
    sheet_name = f"5330_{report_date}{current_time}"  # 工作表名稱
    output_file_name = f"twn_fin_estm_chk_estm_u11_upload_{report_date}_{current_time}.xlsx"  # Excel 檔案名稱
    
    # 設定輸出路徑為與 "python" 資料夾同層級的 "output" 資料夾
    base_path = os.path.dirname(os.getcwd())
    output_dir = os.path.join(base_path, "output")  # 建立 output 資料夾的路徑
    os.makedirs(output_dir, exist_ok=True)  # 如果 output 資料夾不存在，則建立
    output_file = os.path.join(output_dir, output_file_name)  # 完整輸出檔案路徑
    
    # 初始化 Excel 檔案
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # 1~3 列設定固定文字
    ws["A1"] = f"＊Form ID: 5330{' '*100}<報表日期:{report_date} {datetime.now().strftime('%H:%M')}>"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=14)  # 合併儲存格
    ws["A1"].alignment = Alignment(horizontal="left")

    ws["A2"] = "＊Program: /home1/tejdata/report/"
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=14)
    ws["A2"].alignment = Alignment(horizontal="left")

    ws["A3"] = "自結數重大訊息截字結果"
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=14)
    ws["A3"].alignment = Alignment(horizontal="left")

    # 第四列寫入 DataFrame 的欄位名稱
    for col_idx, col_name in enumerate(data_5330.columns, start=1):
        ws.cell(row=4, column=col_idx, value=col_name)

    # 將 DataFrame 資料寫入 Excel（從第 5 列開始）
    for r_idx, row in enumerate(data_5330.itertuples(index=False), start=5):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # 儲存 Excel 檔案
    wb.save(output_file)

    print(f"5330 檔案已成功輸出至: {output_file}")
    